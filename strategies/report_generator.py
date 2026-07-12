# -*- coding: utf-8 -*-
"""
Generador de Reportes de Backtesting en Excel
Genera reportes profesionales en formato .xlsx
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from strategies.backtest_engine import Trade


class ReportGenerator:
    """Genera reportes profesionales de backtesting en Excel"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        # Estilos
        self.header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
        self.title_font = Font(name='Calibri', bold=True, size=14, color='2E4057')
        self.subtitle_font = Font(name='Calibri', bold=True, size=11, color='4A4A4A')
        self.data_font = Font(name='Calibri', size=10)
        self.number_font = Font(name='Calibri', size=10)
        self.positive_font = Font(name='Calibri', size=10, color='008000', bold=True)
        self.negative_font = Font(name='Calibri', size=10, color='CC0000', bold=True)
        self.border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        self.light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    def _apply_header_style(self, ws, row: int, cols: int):
        """Aplica estilo de encabezado a una fila"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
    
    def _apply_data_style(self, ws, row: int, cols: int, is_alternate: bool = False):
        """Aplica estilo a una fila de datos"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.data_font
            cell.border = self.border
            if is_alternate:
                cell.fill = self.light_fill
    
    def _auto_width(self, ws):
        """Ajusta automáticamente el ancho de las columnas"""
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_excel_report(self, backtest_results: Dict) -> str:
        """
        Crea reporte completo en Excel con múltiples hojas
        
        Returns:
            Ruta del archivo Excel generado
        """
        trades = backtest_results['trades']
        ticker_results = backtest_results['ticker_results']
        initial_capital = backtest_results['initial_capital']
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Resumen Ejecutivo
        self._create_summary_sheet(wb, backtest_results)
        
        # Hoja 2: Detalle de Trades
        self._create_trades_sheet(wb, trades)
        
        # Hoja 3: Rendimiento Anual
        self._create_annual_sheet(wb, trades)
        
        # Hoja 4: Métricas por Ticker
        self._create_ticker_sheet(wb, ticker_results)
        
        # Hoja 5: Evolución del Capital
        equity_curve = backtest_results.get('equity_curve', None)
        self._create_equity_curve_sheet(wb, trades, initial_capital, equity_curve)
        
        # Guardar archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'BACKTESTING_REPORT_{timestamp}.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        
        wb.save(filepath)
        print(f"[OK] Reporte Excel guardado: {filepath}")
        
        return filepath
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict):
        """Crea la hoja de resumen ejecutivo"""
        ws = wb.active
        ws.title = "RESUMEN EJECUTIVO"
        
        trades = results['trades']
        initial_capital = results['initial_capital']
        final_capital = results.get('final_capital', initial_capital + sum(t.pnl for t in trades))
        cagr = results.get('cagr', 0)
        total_return = results.get('total_return', ((final_capital - initial_capital) / initial_capital) * 100)
        
        # Calcular metricas
        wins = [t for t in trades if t.pnl > 0]
        losses = [t for t in trades if t.pnl <= 0]
        
        avg_win = np.mean([t.pnl for t in wins]) if wins else 0
        avg_loss = np.mean([t.pnl for t in losses]) if losses else 0
        profit_factor = abs(sum(t.pnl for t in wins) / sum(t.pnl for t in losses)) if losses and sum(t.pnl for t in losses) != 0 else float('inf')
        avg_holding = np.mean([t.holding_days for t in trades]) if trades else 0
        max_drawdown = max(t.max_drawdown for t in trades) if trades else 0
        
        returns = [t.pnl_pct for t in trades]
        sharpe = np.mean(returns) / np.std(returns) if returns and np.std(returns) > 0 else 0
        
        # Titulo
        ws.merge_cells('A1:F1')
        ws['A1'] = "REPORTE RESUMEN EJECUTIVO"
        ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='2E4057')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = "ESTRATEGIA MOMENTUM BREAKOUT - PORTAFOLIO"
        ws['A2'].font = Font(name='Calibri', bold=True, size=12, color='4A4A4A')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Fecha de Generacion: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(name='Calibri', size=10, color='666666')
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Seccion de Capital
        row = 5
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "RESUMEN DE CAPITAL"
        ws[f'A{row}'].font = self.subtitle_font
        ws[f'A{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        capital_data = [
            ("Capital Inicial", f"${initial_capital:,.2f}"),
            ("Capital Final", f"${final_capital:,.2f}"),
            ("Ganancia/Perdida Total", f"${final_capital - initial_capital:,.2f}"),
            ("Retorno Total", f"{total_return:,.2f}%"),
            ("CAGR", f"{cagr:.2f}%"),
        ]
        
        row += 1
        for label, value in capital_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.data_font
            if 'Ganancia' in label or 'Retorno' in label:
                ws[f'B{row}'].font = self.positive_font if (final_capital - initial_capital) > 0 else self.negative_font
            row += 1
        
        # Sección de Métricas
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "MÉTRICAS DE TRADING"
        ws[f'A{row}'].font = self.subtitle_font
        ws[f'A{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        metrics_data = [
            ("Total de Trades", len(trades)),
            ("Trades Ganadores", len(wins)),
            ("Trades Perdedores", len(losses)),
            ("Win Rate", f"{len(wins)/len(trades)*100:.1f}%" if trades else "N/A"),
            ("Ganancia Promedio", f"${avg_win:,.2f}"),
            ("Pérdida Promedio", f"${avg_loss:,.2f}"),
            ("Ratio Promedio", f"{abs(avg_win/avg_loss):.2f}" if avg_loss != 0 else "N/A"),
            ("Profit Factor", f"{profit_factor:.2f}"),
            ("Sharpe Ratio", f"{sharpe:.2f}"),
            ("Días Promedio en Posición", f"{avg_holding:.1f}"),
            ("Max Drawdown", f"{max_drawdown:.2f}%"),
        ]
        
        row += 1
        for label, value in metrics_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.data_font
            row += 1
        
        # Sección de Configuración
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "CONFIGURACIÓN DE LA ESTRATEGIA"
        ws[f'A{row}'].font = self.subtitle_font
        ws[f'A{row}'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        config_data = [
            ("Nombre", "EMA Triple Cross Swing"),
            ("Período", "2016-01-01 a 2025-12-31 (10 años)"),
            ("Temporalidad", "Diaria (Swing Trading)"),
            ("Tipo", "Solo Long (Compra)"),
            ("Indicadores", "EMA 50/100/200 + RSI + MACD + ATR"),
            ("Stop Loss", "Trailing Stop ATR x 3.5"),
            ("Riesgo por Trade", "2% del capital"),
        ]
        
        row += 1
        for label, value in config_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.data_font
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
    
    def _create_trades_sheet(self, wb: Workbook, trades: List[Trade]):
        """Crea la hoja de detalle de trades"""
        ws = wb.create_sheet("DETALLE TRADES")
        
        # Título
        ws.merge_cells('A1:P1')
        ws['A1'] = "DETALLE COMPLETO DE OPERACIONES"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        headers = [
            "#", "TICKER", "CANTIDAD", "FECHA COMPRA", "PRECIO COMPRA",
            "FECHA VENTA", "PRECIO VENTA", "P&L ($)", "P&L (%)",
            "DÍAS", "MAX DD (%)", "MOTIVO COMPRA",
            "MONTO INGRESO", "MONTO SALIDA", "MONTO INVERTIDO", "MONTO TOTAL"
        ]
        
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, len(headers))
        
        # Datos
        row = 4
        for i, trade in enumerate(trades, 1):
            is_alternate = i % 2 == 0
            
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=trade.ticker)
            ws.cell(row=row, column=3, value=trade.quantity)
            ws.cell(row=row, column=4, value=trade.buy_date)
            ws.cell(row=row, column=5, value=trade.buy_price)
            ws.cell(row=row, column=6, value=trade.sell_date)
            ws.cell(row=row, column=7, value=trade.sell_price)
            ws.cell(row=row, column=8, value=trade.pnl)
            ws.cell(row=row, column=9, value=trade.pnl_pct)
            ws.cell(row=row, column=10, value=trade.holding_days)
            ws.cell(row=row, column=11, value=trade.max_drawdown)
            ws.cell(row=row, column=12, value=trade.buy_reason[:100])
            ws.cell(row=row, column=13, value=trade.entry_amount)
            ws.cell(row=row, column=14, value=trade.exit_amount)
            ws.cell(row=row, column=15, value=trade.invested_amount)
            ws.cell(row=row, column=16, value=trade.total_amount)
            
            # Aplicar estilos
            self._apply_data_style(ws, row, len(headers), is_alternate)
            
            # Formato de números
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=9).number_format = '0.00"%"'
            ws.cell(row=row, column=11).number_format = '0.00"%"'
            ws.cell(row=row, column=13).number_format = '#,##0.00'
            ws.cell(row=row, column=14).number_format = '#,##0.00'
            ws.cell(row=row, column=15).number_format = '#,##0.00'
            ws.cell(row=row, column=16).number_format = '#,##0.00'
            
            # Colorear P&L
            pnl_cell = ws.cell(row=row, column=8)
            pnl_pct_cell = ws.cell(row=row, column=9)
            if trade.pnl > 0:
                pnl_cell.font = self.positive_font
                pnl_pct_cell.font = self.positive_font
            else:
                pnl_cell.font = self.negative_font
                pnl_pct_cell.font = self.negative_font
            
            row += 1
        
        # Resumen al final
        row += 1
        ws.cell(row=row, column=1, value="RESUMEN")
        ws.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=11)
        
        total_pnl = sum(t.pnl for t in trades)
        total_trades = len(trades)
        wins = len([t for t in trades if t.pnl > 0])
        total_invested = sum(t.invested_amount for t in trades)
        total_received = sum(t.total_amount for t in trades)
        
        ws.cell(row=row, column=2, value=f"Total Trades: {total_trades}")
        ws.cell(row=row, column=4, value=f"Win Rate: {wins/total_trades*100:.1f}%" if total_trades > 0 else "N/A")
        ws.cell(row=row, column=6, value=f"P&L Total: ${total_pnl:,.2f}")
        ws.cell(row=row, column=13, value=f"Total Invertido: ${total_invested:,.2f}")
        ws.cell(row=row, column=15, value=f"Total Recibido: ${total_received:,.2f}")
        
        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['L'].width = 40
    
    def _create_annual_sheet(self, wb: Workbook, trades: List[Trade]):
        """Crea la hoja de rendimiento anual"""
        ws = wb.create_sheet("RENDIMIENTO ANUAL")
        
        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = "RENDIMIENTO ANUAL POR AÑO"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Agrupar trades por año
        annual_data = {}
        for trade in trades:
            year = trade.buy_date[:4]
            if year not in annual_data:
                annual_data[year] = {
                    'trades': 0, 'wins': 0, 'losses': 0,
                    'total_pnl': 0, 'pnl_list': []
                }
            annual_data[year]['trades'] += 1
            annual_data[year]['total_pnl'] += trade.pnl
            annual_data[year]['pnl_list'].append(trade.pnl)
            if trade.pnl > 0:
                annual_data[year]['wins'] += 1
            else:
                annual_data[year]['losses'] += 1
        
        # Encabezados
        headers = ["AÑO", "TRADES", "WINS", "LOSSES", "WIN RATE", "P&L TOTAL", "P&L PROMEDIO"]
        
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, len(headers))
        
        # Datos
        row = 4
        total_pnl = 0
        total_trades = 0
        total_wins = 0
        
        for year in sorted(annual_data.keys()):
            data = annual_data[year]
            win_rate = (data['wins'] / data['trades'] * 100) if data['trades'] > 0 else 0
            avg_pnl = np.mean(data['pnl_list']) if data['pnl_list'] else 0
            
            is_alternate = (row - 4) % 2 == 1
            
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=data['trades'])
            ws.cell(row=row, column=3, value=data['wins'])
            ws.cell(row=row, column=4, value=data['losses'])
            ws.cell(row=row, column=5, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=6, value=data['total_pnl'])
            ws.cell(row=row, column=7, value=avg_pnl)
            
            self._apply_data_style(ws, row, len(headers), is_alternate)
            
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            
            # Colorear P&L
            pnl_cell = ws.cell(row=row, column=6)
            if data['total_pnl'] > 0:
                pnl_cell.font = self.positive_font
            else:
                pnl_cell.font = self.negative_font
            
            total_pnl += data['total_pnl']
            total_trades += data['trades']
            total_wins += data['wins']
            
            row += 1
        
        # Fila de totales
        overall_win_rate = (total_wins / total_trades * 100) if total_trades > 0 else 0
        
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=total_trades)
        ws.cell(row=row, column=3, value=total_wins)
        ws.cell(row=row, column=4, value=total_trades - total_wins)
        ws.cell(row=row, column=5, value=f"{overall_win_rate:.1f}%")
        ws.cell(row=row, column=6, value=total_pnl)
        ws.cell(row=row, column=7, value=total_pnl / total_trades if total_trades > 0 else 0)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name='Calibri', bold=True, size=10)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = self.border
        
        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_ticker_sheet(self, wb: Workbook, ticker_results: Dict):
        """Crea la hoja de métricas por ticker"""
        ws = wb.create_sheet("MÉTRICAS POR TICKER")
        
        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = "RENDIMIENTO POR ACTIVO"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        headers = [
            "TICKER", "TRADES", "WINS", "LOSSES", "WIN RATE",
            "P&L TOTAL", "P&L PROMEDIO", "MEJOR TRADE", "PEOR TRADE", "RETORNO %"
        ]
        
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, len(headers))
        
        # Datos
        row = 4
        for i, (ticker, results) in enumerate(ticker_results.items()):
            is_alternate = i % 2 == 1
            
            ws.cell(row=row, column=1, value=ticker)
            ws.cell(row=row, column=2, value=results['total_trades'])
            ws.cell(row=row, column=3, value=results['wins'])
            ws.cell(row=row, column=4, value=results['losses'])
            ws.cell(row=row, column=5, value=f"{results['win_rate']:.1f}%")
            ws.cell(row=row, column=6, value=results['total_pnl'])
            ws.cell(row=row, column=7, value=results['avg_pnl'])
            ws.cell(row=row, column=8, value=results['best_trade'])
            ws.cell(row=row, column=9, value=results['worst_trade'])
            ws.cell(row=row, column=10, value=f"{results['return_pct']:.2f}%")
            
            self._apply_data_style(ws, row, len(headers), is_alternate)
            
            # Formato de números
            for col in [6, 7, 8, 9]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            # Colorear P&L
            pnl_cell = ws.cell(row=row, column=6)
            if results['total_pnl'] > 0:
                pnl_cell.font = self.positive_font
            else:
                pnl_cell.font = self.negative_font
            
            row += 1
        
        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_equity_curve_sheet(self, wb: Workbook, trades: List[Trade], initial_capital: float,
                                   equity_curve=None):
        """Crea la hoja de curva de equity"""
        ws = wb.create_sheet("CURVA DE EQUITY")
        
        # Titulo
        ws.merge_cells('A1:E1')
        ws['A1'] = "EVOLUCION DEL CAPITAL"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Si hay datos de equity curve del portafolio, usarlos
        if equity_curve and len(equity_curve) > 0:
            headers = ["FECHA", "CAPITAL EFECTIVO", "VALOR POSICIONES", "VALOR TOTAL", "NUM POSICIONES"]
            
            row = 3
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, len(headers))
            
            row = 4
            for i, data in enumerate(equity_curve):
                is_alternate = i % 2 == 1
                ws.cell(row=row, column=1, value=data['date'].strftime('%Y-%m-%d') if hasattr(data['date'], 'strftime') else str(data['date']))
                ws.cell(row=row, column=2, value=data['capital'])
                ws.cell(row=row, column=3, value=data['positions_value'])
                ws.cell(row=row, column=4, value=data['total_value'])
                ws.cell(row=row, column=5, value=data['num_positions'])
                
                self._apply_data_style(ws, row, len(headers), is_alternate)
                
                for col in [2, 3, 4]:
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
                
                row += 1
        else:
            # Fallback: usar trades
            equity_data = [{'date': 'Inicio', 'capital': initial_capital, 'trade': '-', 'pnl': 0}]
            capital = initial_capital
            for trade in trades:
                capital += trade.pnl
                equity_data.append({
                    'date': trade.sell_date,
                    'capital': capital,
                    'trade': f"{trade.ticker} ({trade.pnl_pct:+.2f}%)",
                    'pnl': trade.pnl
                })
            
            headers = ["FECHA", "CAPITAL ACUMULADO", "TRADE", "P&L TRADE"]
            row = 3
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, len(headers))
            
            row = 4
            for i, data in enumerate(equity_data):
                is_alternate = i % 2 == 1
                ws.cell(row=row, column=1, value=data['date'])
                ws.cell(row=row, column=2, value=data['capital'])
                ws.cell(row=row, column=3, value=data['trade'])
                ws.cell(row=row, column=4, value=data['pnl'])
                self._apply_data_style(ws, row, len(headers), is_alternate)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                row += 1
        
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
    
    def save_reports(self, backtest_results: Dict):
        """Guarda todos los reportes"""
        # Generar Excel
        excel_path = self.create_excel_report(backtest_results)
        
        # También guardar CSV para análisis adicional
        trades = backtest_results['trades']
        if trades:
            trades_df = pd.DataFrame([t.to_dict() for t in trades])
            csv_path = os.path.join(self.output_dir, 'trades.csv')
            trades_df.to_csv(csv_path, index=False)
            print(f"[OK] CSV guardado: {csv_path}")
        
        print(f"\n[OK] Todos los reportes guardados en: {self.output_dir}")
        
        return excel_path
