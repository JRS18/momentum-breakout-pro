# -*- coding: utf-8 -*-
"""
Tracker de Operaciones Reales - Momentum Breakout Pro
Registra compras/ventas reales y calcula capital disponible
"""
import json, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RUTA = os.path.dirname(os.path.abspath(__file__))
OPERACIONES_PATH = os.path.join(RUTA, 'operaciones.json')
TRACKER_PATH = os.path.join(RUTA, 'tracker_operaciones.xlsx')

CEDEAR_RATIOS = {
    "NBIS": 27, "RGTI": 2, "IREN": 12, "RIOT": 3, "HUT": 5, "UPST": 5, "ALAB": 44, "PLTR": 3
}


def cargar_operaciones():
    if os.path.exists(OPERACIONES_PATH):
        with open(OPERACIONES_PATH) as f:
            return json.load(f)
    return {'operaciones': [], 'capital_inicial': 5000, 'comision': 0.006, 'impuesto': 0.015}


def guardar_operaciones(data):
    with open(OPERACIONES_PATH, 'w') as f:
        json.dump(data, f, indent=2, default=str)


def registrar_compra(ticker, fecha, cedears, precio_cedear_ars, ccl):
    data = cargar_operaciones()
    ratio = CEDEAR_RATIOS.get(ticker, 1)
    precio_usd = precio_cedear_ars * ratio / ccl
    costo_ars = cedears * precio_cedear_ars
    comision_ars = costo_ars * data['comision']
    total_ars = costo_ars + comision_ars

    op = {
        'tipo': 'COMPRA',
        'ticker': ticker,
        'fecha': fecha,
        'ratio': ratio,
        'cedears': cedears,
        'precio_cedear_ars': precio_cedear_ars,
        'precio_usd': precio_usd,
        'ccl': ccl,
        'costo_ars': costo_ars,
        'comision_ars': comision_ars,
        'total_ars': total_ars
    }
    data['operaciones'].append(op)
    guardar_operaciones(data)
    return op


def registrar_venta(ticker, fecha, cedears, precio_cedear_ars, ccl, precio_compra_ars=None):
    data = cargar_operaciones()
    ratio = CEDEAR_RATIOS.get(ticker, 1)
    precio_usd = precio_cedear_ars * ratio / ccl
    ingreso_ars = cedears * precio_cedear_ars
    comision_ars = ingreso_ars * data['comision']

    # Buscar precio de compra para calcular ganancia
    if precio_compra_ars is None:
        for op in reversed(data['operaciones']):
            if op['ticker'] == ticker and op['tipo'] == 'COMPRA':
                precio_compra_ars = op['precio_cedear_ars']
                break

    ganancia_ars = 0
    impuesto_ars = 0
    if precio_compra_ars:
        ganancia_ars = (precio_cedear_ars - precio_compra_ars) * cedears
        if ganancia_ars > 0:
            impuesto_ars = ganancia_ars * data['impuesto']

    neto_ars = ingreso_ars - comision_ars - impuesto_ars

    op = {
        'tipo': 'VENTA',
        'ticker': ticker,
        'fecha': fecha,
        'ratio': ratio,
        'cedears': cedears,
        'precio_cedear_ars': precio_cedear_ars,
        'precio_usd': precio_usd,
        'ccl': ccl,
        'ingreso_ars': ingreso_ars,
        'comision_ars': comision_ars,
        'ganancia_ars': ganancia_ars,
        'impuesto_ars': impuesto_ars,
        'neto_ars': neto_ars
    }
    data['operaciones'].append(op)
    guardar_operaciones(data)
    return op


def calcular_estado():
    data = cargar_operaciones()
    # Convertir capital inicial a ARS
    try:
        import requests
        r = requests.get('https://dolarapi.com/v1/dolares/ccl', timeout=5)
        ccl = r.json().get('venta', 1550)
    except:
        ccl = 1550
    capital = data['capital_inicial'] * ccl  # Convertir USD a ARS
    posiciones = {}

    for op in data['operaciones']:
        if op['tipo'] == 'COMPRA':
            capital -= op['total_ars']
            if op['ticker'] not in posiciones:
                posiciones[op['ticker']] = {'cedears': 0, 'costo_total': 0}
            posiciones[op['ticker']]['cedears'] += op['cedears']
            posiciones[op['ticker']]['costo_total'] += op['total_ars']
        elif op['tipo'] == 'VENTA':
            capital += op['neto_ars']
            if op['ticker'] in posiciones:
                posiciones[op['ticker']]['cedears'] -= op['cedears']
                if posiciones[op['ticker']]['cedears'] <= 0:
                    del posiciones[op['ticker']]

    return {
        'capital_inicial': data['capital_inicial'],
        'capital_inicial_ars': data['capital_inicial'] * ccl,
        'capital_disponible_ars': capital,
        'posiciones': posiciones,
        'operaciones': data['operaciones']
    }


def fmt_ars(valor):
    return f"{valor:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")


def generar_tracker_excel():
    estado = calcular_estado()
    wb = Workbook()

    # ========== HOJA 1: Resumen ==========
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = "74B9FF"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    title_font = Font(bold=True, size=14, color="0F3460")

    ws1.merge_cells('A1:F1')
    ws1['A1'] = "RESUMEN DE OPERACIONES"
    ws1['A1'].font = title_font

    ws1['A3'] = "Capital Inicial"
    ws1['B3'] = f"USD {estado['capital_inicial']:,.2f}"
    ws1['A4'] = "Capital Disponible"
    ws1['B4'] = f"ARS {fmt_ars(estado['capital_disponible_ars'])}"
    ws1['A5'] = "Posiciones Activas"
    ws1['B5'] = len(estado['posiciones'])

    total_ops = len(estado['operaciones'])
    compras = [o for o in estado['operaciones'] if o['tipo'] == 'COMPRA']
    ventas = [o for o in estado['operaciones'] if o['tipo'] == 'VENTA']
    ws1['A7'] = "Total Operaciones"
    ws1['B7'] = total_ops
    ws1['A8'] = "Compras"
    ws1['B8'] = len(compras)
    ws1['A9'] = "Ventas"
    ws1['B9'] = len(ventas)

    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ========== HOJA 2: Posiciones Activas ==========
    ws2 = wb.create_sheet("Posiciones")
    ws2.sheet_properties.tabColor = "FDCB6E"

    ws2.merge_cells('A1:F1')
    ws2['A1'] = "POSICIONES ACTIVAS"
    ws2['A1'].font = title_font

    headers = ["Ticker", "CEDEARs", "Ratio", "P. Compra (ARS)", "Invertido (ARS)", "P. Actual (ARS)"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 4
    for ticker, pos in estado['posiciones'].items():
        ratio = CEDEAR_RATIOS.get(ticker, 1)
        promedio = pos['costo_total'] / pos['cedears'] if pos['cedears'] > 0 else 0
        ws2.cell(row=row, column=1, value=ticker).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=pos['cedears'])
        ws2.cell(row=row, column=3, value=f"{ratio}:1")
        ws2.cell(row=row, column=4, value=f"ARS {fmt_ars(promedio)}")
        ws2.cell(row=row, column=5, value=f"ARS {fmt_ars(pos['costo_total'])}")
        ws2.cell(row=row, column=6, value="ARS -")
        row += 1

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ========== HOJA 3: Historial ==========
    ws3 = wb.create_sheet("Historial")
    ws3.sheet_properties.tabColor = "00B894"

    ws3.merge_cells('A1:J1')
    ws3['A1'] = "HISTORIAL DE OPERACIONES"
    ws3['A1'].font = title_font

    headers = ["Fecha", "Tipo", "Ticker", "Ratio", "CEDEARs", "P. CEDEAR (ARS)",
               "Monto (ARS)", "Comision (ARS)", "Impuesto (ARS)", "Neto (ARS)"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 4
    for op in estado['operaciones']:
        if op['tipo'] == 'COMPRA':
            ws3.cell(row=row, column=1, value=op['fecha'])
            ws3.cell(row=row, column=2, value="COMPRA").font = Font(color="00B894", bold=True)
            ws3.cell(row=row, column=3, value=op['ticker']).font = Font(bold=True)
            ws3.cell(row=row, column=4, value=f"{op['ratio']}:1")
            ws3.cell(row=row, column=5, value=op['cedears'])
            ws3.cell(row=row, column=6, value=f"ARS {fmt_ars(op['precio_cedear_ars'])}")
            ws3.cell(row=row, column=7, value=f"ARS {fmt_ars(op['costo_ars'])}")
            ws3.cell(row=row, column=8, value=f"ARS {fmt_ars(op['comision_ars'])}")
            ws3.cell(row=row, column=9, value="ARS 0")
            ws3.cell(row=row, column=10, value=f"-ARS {fmt_ars(op['total_ars'])}")
        else:
            ws3.cell(row=row, column=1, value=op['fecha'])
            ws3.cell(row=row, column=2, value="VENTA").font = Font(color="E17055", bold=True)
            ws3.cell(row=row, column=3, value=op['ticker']).font = Font(bold=True)
            ws3.cell(row=row, column=4, value=f"{op['ratio']}:1")
            ws3.cell(row=row, column=5, value=op['cedears'])
            ws3.cell(row=row, column=6, value=f"ARS {fmt_ars(op['precio_cedear_ars'])}")
            ws3.cell(row=row, column=7, value=f"ARS {fmt_ars(op['ingreso_ars'])}")
            ws3.cell(row=row, column=8, value=f"ARS {fmt_ars(op['comision_ars'])}")
            ws3.cell(row=row, column=9, value=f"ARS {fmt_ars(op['impuesto_ars'])}")
            ws3.cell(row=row, column=10, value=f"ARS {fmt_ars(op['neto_ars'])}")
        row += 1

    for col in range(1, 11):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    # ========== HOJA 4: Capital ==========
    ws4 = wb.create_sheet("Capital")
    ws4.sheet_properties.tabColor = "E17055"

    ws4.merge_cells('A1:D1')
    ws4['A1'] = "EVOLUCION DE CAPITAL"
    ws4['A1'].font = title_font

    headers = ["Fecha", "Concepto", "Monto (ARS)", "Saldo (ARS)"]
    for col, h in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    saldo = estado['capital_inicial']
    # Convertir capital inicial a ARS (aprox)
    saldo_ars = saldo * 1530  # MEP aprox
    ws4.cell(row=4, column=1, value="Inicio")
    ws4.cell(row=4, column=2, value="Capital Inicial")
    ws4.cell(row=4, column=3, value=f"USD {fmt_ars(saldo)}")
    ws4.cell(row=4, column=4, value=f"ARS {fmt_ars(saldo_ars)}")

    row = 5
    for op in estado['operaciones']:
        if op['tipo'] == 'COMPRA':
            saldo_ars -= op['total_ars']
            ws4.cell(row=row, column=1, value=op['fecha'])
            ws4.cell(row=row, column=2, value=f"Compra {op['ticker']}")
            ws4.cell(row=row, column=3, value=f"-ARS {fmt_ars(op['total_ars'])}")
            ws4.cell(row=row, column=4, value=f"ARS {fmt_ars(saldo_ars)}")
        else:
            saldo_ars += op['neto_ars']
            ws4.cell(row=row, column=1, value=op['fecha'])
            ws4.cell(row=row, column=2, value=f"Venta {op['ticker']}")
            ws4.cell(row=row, column=3, value=f"+ARS {fmt_ars(op['neto_ars'])}")
            ws4.cell(row=row, column=4, value=f"ARS {fmt_ars(saldo_ars)}")
        row += 1

    for col in range(1, 5):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    wb.save(TRACKER_PATH)
    print(f"  [OK] Tracker guardado: {TRACKER_PATH}")
    return TRACKER_PATH


if __name__ == '__main__':
    generar_tracker_excel()
